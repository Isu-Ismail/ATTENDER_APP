import os
import sys
import openpyxl as xl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, datetime
import customtkinter as ctk
from tkinter import messagebox
from config import ICON_PATH, USER_DATA_PATH, resource_path
from excel_helpers import count_student_rows
from ui_windows import LowAttendanceWindow, ManageWindow, DetailedReportWindow,BulkEntryWindow, MarkEntryWindow, LiveSessionWindow
import requests
import threading

# --- Main Application Class ---
class AttendanceApp(ctk.CTk):
    """The main application class."""
    def __init__(self):
        super().__init__()
        
        # --- 1. Basic Window Setup ---
        self.title("Attendance Marker")
        self.geometry("500x700")
        self.resizable(False, False)
        try:
            self.iconbitmap(resource_path(ICON_PATH))
        except Exception as e:
            print(f"Icon not found at '{ICON_PATH}'. Skipping. Error: {e}")
        
        # --- 2. Grid Configuration ---
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1) # Makes the scrollable frame expand
        
        # --- 3. Initialize All Instance Variables ---
        # Initialize pop-up window trackers
        self.manage_win = self.report_win = self.detail_win = self.bulk_win = self.mark_win = None
        # Initialize file and data variables
        self.current_filename = None
        self.wb = None
        
        # --- THIS IS THE FIX ---
        # Initialize all widget variables to None to prevent AttributeErrors
        self.file_combo = self.open_button = self.load_button = None
        self.subject_combo = self.date_entry = self.hours_entry = None
        self.mode_var = ctk.StringVar(value="absent") # This one needs to be created
        self.absent_btn = self.present_btn = self.rolls_entry = None
        self.submit_button = self.report_button = self.detailed_report_button = None
        self.bulk_entry_button = self.mark_entry_button = self.manage_button = None
        self.live_session_button = self.status_frame = self.status_label = None

        # --- 4. Build UI and Set Initial State ---
        self.setup_ui()
        self.set_main_controls_state("disabled")
    
    def setup_ui(self):
        # --- Top frame (not scrollable) ---
        file_frame = ctk.CTkFrame(self)
        file_frame.grid(row=0, column=0, padx=20, pady=(20, 10), sticky="ew")
        file_frame.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(file_frame, text="Attendance File:", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, columnspan=3, padx=10, pady=(10,0), sticky="w")
        self.file_combo = ctk.CTkComboBox(file_frame, values=self.find_excel_files(), command=self.file_selected)
        self.file_combo.grid(row=1, column=0, padx=(10,5), pady=10, sticky="ew")
        self.file_combo.bind("<FocusIn>", self.clear_file_combo_placeholder)
        self.file_combo.set("Select a file or type a new name")
        self.open_button = ctk.CTkButton(file_frame, text="Open File", width=100, command=self.open_selected_file)
        self.open_button.grid(row=1, column=1, padx=5, pady=10)
        self.load_button = ctk.CTkButton(file_frame, text="Load File", width=100, command=self.load_file)
        self.load_button.grid(row=1, column=2, padx=(0,10), pady=10)
        
        # --- NEW: Main scrollable frame for all other content ---
        content_frame = ctk.CTkScrollableFrame(self)
        content_frame.grid(row=1, column=0, padx=20, pady=0, sticky="nsew")
        content_frame.grid_columnconfigure(0, weight=1)

        # --- All widgets below are now placed inside 'content_frame' ---
        self.main_frame = ctk.CTkFrame(content_frame)
        self.main_frame.grid(row=0, column=0, pady=10, sticky="ew")
        self.main_frame.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(self.main_frame, text="Select Subject", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, columnspan=2, padx=10, pady=(10, 5), sticky="w")
        self.subject_combo = ctk.CTkComboBox(self.main_frame, state="readonly", values=[])
        self.subject_combo.grid(row=1, column=0, columnspan=2, padx=10, pady=(0, 10), sticky="ew")
        ctk.CTkLabel(self.main_frame, text="Date for Session", font=ctk.CTkFont(weight="bold")).grid(row=2, column=0, padx=10, pady=(10, 5), sticky="w")
        self.date_entry = ctk.CTkEntry(self.main_frame, placeholder_text="DD-MM-YYYY")
        self.date_entry.grid(row=3, column=0, padx=10, pady=(0, 10), sticky="ew")
        self.date_entry.insert(0, date.today().strftime("%d-%m-%Y"))
        ctk.CTkLabel(self.main_frame, text="Hours for Session", font=ctk.CTkFont(weight="bold")).grid(row=2, column=1, padx=10, pady=(10, 5), sticky="w")
        self.hours_entry = ctk.CTkEntry(self.main_frame, placeholder_text="e.g., 2")
        self.hours_entry.grid(row=3, column=1, padx=10, pady=(0, 10), sticky="ew")
        ctk.CTkLabel(self.main_frame, text="Mark by listing:", font=ctk.CTkFont(weight="bold")).grid(row=4, column=0, columnspan=2, padx=10, pady=(10, 5), sticky="w")
        self.mode_var = ctk.StringVar(value="absent")
        mode_frame = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        mode_frame.grid(row=5, column=0, columnspan=2, padx=10, pady=(0, 5), sticky="ew")
        self.absent_btn = ctk.CTkRadioButton(mode_frame, text="Absentees", variable=self.mode_var, value="absent")
        self.absent_btn.pack(side="left")
        self.present_btn = ctk.CTkRadioButton(mode_frame, text="Presentees", variable=self.mode_var, value="present")
        self.present_btn.pack(side="left", padx=20)
        ctk.CTkLabel(self.main_frame, text="Enter Roll Numbers (comma-separated)", font=ctk.CTkFont(weight="bold")).grid(row=6, column=0, columnspan=2, padx=10, pady=(10, 5), sticky="w")
        self.rolls_entry = ctk.CTkEntry(self.main_frame, placeholder_text="e.g., 2, 4, 9")
        self.rolls_entry.grid(row=7, column=0, columnspan=2, padx=10, pady=(0, 20), sticky="ew")

        self.submit_button = ctk.CTkButton(content_frame, text="Mark Attendance", command=self.validate_and_submit)
        self.submit_button.grid(row=1, column=0, pady=(10, 5), sticky="ew")
        
        tools_frame1 = ctk.CTkFrame(content_frame, fg_color="transparent")
        tools_frame1.grid(row=2, column=0, pady=5, sticky="ew")
        tools_frame1.grid_columnconfigure((0,1), weight=1)
        self.report_button = ctk.CTkButton(tools_frame1, text="Low Attendance", fg_color="#D35400", hover_color="#E67E22", command=self.open_low_attendance_window)
        self.report_button.grid(row=0, column=0, padx=(0,5), sticky="ew")
        self.detailed_report_button = ctk.CTkButton(tools_frame1, text="Detailed Report", command=self.open_detailed_report_window)
        self.detailed_report_button.grid(row=0, column=1, padx=(5,0), sticky="ew")
        
        tools_frame2 = ctk.CTkFrame(content_frame, fg_color="transparent")
        tools_frame2.grid(row=3, column=0, pady=5, sticky="ew")
        tools_frame2.grid_columnconfigure((0,1), weight=1)
        self.bulk_entry_button = ctk.CTkButton(tools_frame2, text="Bulk Entry", command=self.open_bulk_entry_window)
        self.bulk_entry_button.grid(row=0, column=0, padx=(0,5), sticky="ew")
        self.mark_entry_button = ctk.CTkButton(tools_frame2, text="Mark Entry", command=self.open_mark_entry_window)
        self.mark_entry_button.grid(row=0, column=1, padx=(5,0), sticky="ew")
        
        self.manage_button = ctk.CTkButton(content_frame, text="Manage Subjects & Students", command=self.open_manage_window)
        self.manage_button.grid(row=4, column=0, pady=5, sticky="ew")

        self.live_session_button = ctk.CTkButton(content_frame, text="Start Live OTP Session", fg_color="green", hover_color="#006400", command=self.open_live_session_window)
        self.live_session_button.grid(row=5, column=0, pady=(5,10), sticky="ew")

        # --- Status frame (not scrollable) ---
        self.status_frame = ctk.CTkFrame(self)
        self.status_frame.grid(row=2, column=0, padx=20, pady=(10, 10), sticky="ew")
        self.status_label = ctk.CTkLabel(self.status_frame, text="", wraplength=450)
        self.status_label.pack(fill="both", expand=True, padx=5, pady=5)
    
    def open_bulk_entry_window(self):
        """Opens the new bulk entry window."""
        self.hide_status()
        if not self.wb: return self.show_status("No file loaded.", is_error=True)
        subject_name = self.subject_combo.get()
        if not subject_name: return self.show_status("Please select a subject first.", is_error=True)
        
        if self.bulk_win is not None and self.bulk_win.winfo_exists():
            return self.bulk_win.focus()

        
        try:
            sheet = self.wb[subject_name]
            self.bulk_win = BulkEntryWindow(self, sheet)
        except Exception as e:
            self.show_status(f"Could not open Bulk Entry window: {e}", is_error=True)

    def get_student_list(self, sheet):
        """Gets a list of all student names from the sheet."""
        return [str(sheet.cell(row=row, column=2).value) for row in range(5, count_student_rows(sheet) + 5) if sheet.cell(row=row, column=2).value]

    def _find_percentage_col(self, sheet):
        """Helper to find the last PERCENTAGE column."""
        for col in range(sheet.max_column, 3, -1):
            if sheet.cell(row=4, column=col).value == "PERCENTAGE":
                return col
        return None

    def get_assessment_list(self, sheet):
        """Finds all assessment columns (those after the fixed summary block)."""
        assessments = []
        # Assessments start after the fixed summary block (after column Z=26)
        for col in range(27, self._find_true_last_column(sheet) + 2):
            header = sheet.cell(row=4, column=col).value
            max_mark_header = sheet.cell(row=3, column=col).value
            if header and max_mark_header:
                assessments.append(header.strip())
        return sorted(list(set(assessments)))

    def get_marks_for_assessment(self, sheet, assessment_name):
        """Gets a list of marks for a given assessment column."""
        col_idx = None
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=4, column=col).value == assessment_name:
                col_idx = col
                break
        if not col_idx: return []
        
        num_students = count_student_rows(sheet)
        return [str(sheet.cell(row, col_idx).value or '') for row in range(5, num_students + 5)]

    def add_new_assessment_column(self, sheet, name, max_marks):
        """Adds a new assessment column and safely removes any old final result column."""
        try:
            int(max_marks)
        except (ValueError, TypeError):
            return False, "Maximum Marks must be a number."
        
        new_name_upper = name.strip().upper()
        if new_name_upper in [a.upper() for a in self.get_assessment_list(sheet)]:
            return False, f"An assessment named '{name}' already exists."

        # --- NEW: Specifically find and delete the "FINAL RESULT" column ---
        final_result_col = None
        for col in range(sheet.max_column, 3, -1):
            if sheet.cell(row=4, column=col).value == "FINAL RESULT":
                final_result_col = col
                break
        
        if final_result_col:
            if messagebox.askyesno("Update Detected", "An old 'FINAL RESULT' column was found. It is now outdated and will be removed.\n\nYou will need to run the calculator again after entering marks.\n\nProceed?"):
                sheet.delete_cols(final_result_col)
            else:
                return False, "Operation cancelled by user."
        
        new_col = self._find_true_last_column(sheet) + 1
        
        sheet.cell(row=3, column=new_col).value = f"Out of: {max_marks}"
        sheet.cell(row=4, column=new_col).value = name.upper()
        header_font = Font(bold=True, name='Calibri', color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        sheet.cell(row=4, column=new_col).font = header_font
        sheet.cell(row=4, column=new_col).fill = header_fill

        self.apply_standard_styles(sheet, count_student_rows(sheet))
        self.wb.save(os.path.join(USER_DATA_PATH, self.current_filename))
        return True, f"Assessment '{name}' added successfully."

    def save_marks(self, sheet, assessment_name, marks_list):
        """Saves a list of integer marks to the specified assessment column."""
        col_idx = None
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=4, column=col).value == assessment_name:
                col_idx = col
                break
        if not col_idx: return False, "Could not find the assessment column."

        try:
            for i, mark in enumerate(marks_list):
                sheet.cell(row=i + 5, column=col_idx).value = mark
            
            self.wb.save(os.path.join(USER_DATA_PATH, self.current_filename))
            return True, f"Marks for '{assessment_name}' saved successfully."
        except Exception as e:
            return False, f"An error occurred while saving: {e}"

    def get_max_marks(self, sheet, assessment_name):
        """Finds the 'Out of: XX' value for a given assessment."""
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=4, column=col).value == assessment_name:
                max_mark_str = str(sheet.cell(row=3, column=col).value or '').replace('Out of: ', '')
                try:
                    return int(max_mark_str)
                except: return None
        return None
    
    def get_all_dates_from_sheet(self, sheet):
        """Gets a list of all unique attendance dates from the sheet."""
        dates = []
        for col in range(4, sheet.max_column + 1):
            date_val = sheet.cell(row=2, column=col).value
            if isinstance(date_val, str) and '-' in date_val:
                if date_val not in dates:
                    dates.append(date_val)
        return dates
    
    def clear_file_combo_placeholder(self, event):
        if self.file_combo.get() == "Select a file or type a new name":
            self.file_combo.set("")
    
    def find_excel_files(self): 
        return [f for f in os.listdir(USER_DATA_PATH) if f.endswith('.xlsx')]
        
    def file_selected(self, choice): 
        self.current_filename = choice

    def hide_status(self): 
        self.status_frame.grid_forget()

    def load_file(self):
        self.hide_status()
        filename = self.file_combo.get()
        if not filename or "Select" in filename: return self.show_status("Please select or enter a filename.", is_error=True)
        if not filename.endswith('.xlsx'): filename += '.xlsx'
        self.current_filename = filename
        full_path = os.path.join(USER_DATA_PATH, self.current_filename)
        try:
            self.wb = xl.load_workbook(full_path)
            self.show_status(f"Successfully loaded '{self.current_filename}'.")
            self.set_main_controls_state("normal")
            self.update_main_subject_list()
        except FileNotFoundError:
            self.wb = None
            self.show_status(f"File '{self.current_filename}' not found. Use 'Manage' to create it.", is_error=True)
            self.set_main_controls_state("disabled", allow_manage=True)
        except Exception as e:
            self.show_status(f"Error loading file: {e}", is_error=True)
            self.set_main_controls_state("disabled")

    def open_selected_file(self):
        self.hide_status()
        filename = self.file_combo.get()
        if not filename or "Select a file" in filename: return self.show_status("Please select a file to open.", is_error=True)
        if not filename.endswith('.xlsx'): filename += '.xlsx'
        full_path = os.path.join(USER_DATA_PATH, filename)
        if not os.path.exists(full_path): return self.show_status(f"File '{filename}' does not exist.", is_error=True)
        try:
            os.startfile(full_path)
            self.show_status(f"Opening '{filename}'...")
        except Exception as e:
            self.show_status(f"Could not open file: {e}", is_error=True)

    def convert_marks(self, sheet, assessment_name, current_max, new_max):
        """Converts all marks in a column from one scale to another."""
        col_idx = None
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=4, column=col).value == assessment_name:
                col_idx = col
                break
        if not col_idx: return False, "Could not find assessment column."
        
        try:
            num_students = count_student_rows(sheet)
            for row in range(5, num_students + 5):
                cell = sheet.cell(row=row, column=col_idx)
                if cell.value is not None:
                    old_mark = int(cell.value)
                    # Perform conversion and round to nearest whole number
                    new_mark = round((old_mark / current_max) * new_max)
                    cell.value = new_mark
            
            # Update the max mark header
            sheet.cell(row=3, column=col_idx).value = f"Out of: {new_max}"
            self.wb.save(os.path.join(USER_DATA_PATH, self.current_filename))
            return True, "Marks converted successfully."
        except Exception as e:
            return False, f"An error occurred during conversion: {e}"

    def calculate_final_result(self, sheet, weights_dict, final_col_name):
        """Calculates a weighted final score and adds it to a new, styled column."""
        try:
            # First, delete any pre-existing "FINAL RESULT" column to ensure a clean slate
            final_result_col = None
            for col in range(sheet.max_column, 3, -1):
                if sheet.cell(row=4, column=col).value == "FINAL RESULT":
                    final_result_col = col
                    break
            if final_result_col:
                sheet.delete_cols(final_result_col)

            assessment_data = {}
            for name in weights_dict.keys():
                max_mark = self.get_max_marks(sheet, name)
                col_idx = [c for c in range(1, sheet.max_column + 1) if sheet.cell(row=4, column=c).value == name][0]
                if max_mark is None or col_idx is None: return False, f"Could not find data for '{name}'."
                assessment_data[name] = {'col': col_idx, 'max': max_mark}

            new_col_idx = self._find_true_last_column(sheet) + 1
            final_header_cell = sheet.cell(row=4, column=new_col_idx)
            final_header_cell.value = final_col_name.upper()
            header_font = Font(bold=True, name='Calibri', color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            final_header_cell.font, final_header_cell.fill = header_font, header_fill
            
            num_students = count_student_rows(sheet)
            for row in range(5, num_students + 5):
                final_score = 0.0
                for name, weight in weights_dict.items():
                    data = assessment_data[name]
                    student_mark = sheet.cell(row=row, column=data['col']).value or 0
                    contribution = (float(student_mark) / data['max']) * weight
                    final_score += contribution
                sheet.cell(row=row, column=new_col_idx).value = f"{final_score:.2f}"
            
            self.apply_standard_styles(sheet, num_students)
            self.wb.save(os.path.join(USER_DATA_PATH, self.current_filename))
            return True, "Final result calculated successfully."
        except Exception as e:
            return False, f"An error occurred during calculation: {e}"

    def show_status(self, message, is_error=False):
        colors = ("#D5E8D4", "#2E4B2E", "#1E601E", "#90EE90", "✅") if not is_error else ("#FFD2D2", "#5E2D2D", "#C00000", "#FF8282", "❌")
        self.status_frame.configure(fg_color=(colors[0], colors[1]))
        self.status_label.configure(text=f"{colors[4]} {message}", text_color=(colors[2], colors[3]))
        self.status_frame.grid(row=6, column=0, padx=20, pady=(10, 10), sticky="ew")
        
    def set_main_controls_state(self, state="normal", allow_manage=False):
        widgets = [self.subject_combo, self.date_entry, self.hours_entry, self.rolls_entry, 
                   self.submit_button, self.report_button, self.detailed_report_button, 
                   self.bulk_entry_button, self.mark_entry_button, self.absent_btn, 
                   self.present_btn, self.live_session_button]
        for widget in widgets: widget.configure(state=state)
        self.manage_button.configure(state="normal" if state == "normal" or allow_manage else "disabled")
    
    def open_manage_window(self):
        if not self.current_filename: return self.show_status("Please load or name a file first.", is_error=True)
        if self.manage_win and self.manage_win.winfo_exists(): return self.manage_win.focus()
        self.manage_win = ManageWindow(self)

    def update_main_subject_list(self):
        if self.wb: self.subject_combo.configure(values=self.wb.sheetnames)
        else: self.subject_combo.configure(values=[])
        self.subject_combo.set('' if not (self.wb and self.wb.sheetnames) else self.wb.sheetnames[0])
            
    def format_new_sheet(self, sheet):
        """Applies all standard headers, including a FIXED summary block, to a new worksheet."""
        sheet.sheet_view.showGridLines = False
        title_font = Font(size=18, bold=True, name='Calibri')
        header_font = Font(bold=True, name='Calibri', color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        # Main Subject Title
        title_cell = sheet['D1']
        title_cell.value = sheet.title.upper()
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Static Labels
        sheet['B2'].value, sheet['B3'].value = "DATE :", "Hours Taken :"
        for cell_ref in ['B2', 'B3']: sheet[cell_ref].font = Font(bold=True)
        
        # Main table headers for student info
        main_headers = {'A4': 'ROLL NO.', 'B4': 'NAME', 'C4': 'ROLL NUMBER'}
        for cell_ref, text in main_headers.items():
            cell = sheet[cell_ref]
            cell.value, cell.font, cell.fill = text, header_font, header_fill
        
        # --- Create the FIXED summary headers starting at Column W ---
        summary_headers = {
            'W4': 'TOTAL HOURS', 'X4': 'HOURS PRESENT',
            'Y4': 'HOURS ABSENT', 'Z4': 'PERCENTAGE'
        }
        for cell_ref, text in summary_headers.items():
            cell = sheet[cell_ref]
            cell.value, cell.font, cell.fill = text, header_font, header_fill

        self.apply_standard_styles(sheet, 0)

    def apply_standard_styles(self, sheet, num_students):
        """Applies all standard styling: alignment, borders, and column widths."""
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin_side = Side(border_style="thin", color="000000")
        full_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        last_col = self._find_true_last_column(sheet)
        
        # --- UPDATED COLUMN WIDTHS ---
        # Set fixed widths for key areas
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 35
        sheet.column_dimensions['C'].width = 15
        # Set specific width for the summary block
        for col_letter in ['W', 'X', 'Y', 'Z']:
            sheet.column_dimensions[col_letter].width = 15
            
        # Set all other columns (attendance, marks) to a width of 18
        if last_col > 3:
            for col_idx in range(4, last_col + 1):
                col_letter = get_column_letter(col_idx)
                # Only set width if it's not one of the special summary columns
                if col_letter not in ['W', 'X', 'Y', 'Z']:
                    sheet.column_dimensions[col_letter].width = 15

        # Apply borders and alignment to the entire data area
        for row_idx in range(1, num_students + 5):
            for col_idx in range(1, last_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.border = full_border
                cell.alignment = center_align
        
        # Re-apply specific left-alignment for the student names
        for row_idx in range(5, num_students + 5):
            sheet.cell(row=row_idx, column=2).alignment = left_align
 
    def _find_true_last_column(self, sheet):
        """Calculates the last column that contains actual header data by scanning backwards."""
        for col in range(256, 3, -1):
            if (sheet.cell(row=1, column=col).value or
                sheet.cell(row=3, column=col).value or
                sheet.cell(row=4, column=col).value):
                return col
        return 26 

    def _find_percentage_col(self, sheet):
        """Dynamically finds the column number for the 'PERCENTAGE' summary header."""
        # Scan backwards from the last column to find the header in row 4
        for col in range(sheet.max_column, 3, -1): # Scans from right-to-left
            if sheet.cell(row=4, column=col).value == "PERCENTAGE":
                return col # Return the column number as soon as it's found
        return None # Return None if the header is not found for any reason

    def open_mark_entry_window(self):
        """Opens the new mark entry window and assigns it to self.mark_win."""
        self.hide_status()
        if not self.wb: 
            return self.show_status("No file loaded.", is_error=True)
        subject_name = self.subject_combo.get()
        if not subject_name: 
            return self.show_status("Please select a subject first.", is_error=True)
        
        # --- THIS IS THE FIX ---
        # This now correctly checks if the window object exists before using it
        if self.mark_win is not None and self.mark_win.winfo_exists():
            return self.mark_win.focus()
            
        try:
            sheet = self.wb[subject_name]
            self.mark_win = MarkEntryWindow(self, sheet)
        except Exception as e:
            self.show_status(f"Could not open Mark Entry window: {e}", is_error=True)

    def get_report_by_date(self, sheet, dates_list):
        """Generates a summary of attendance for a list of dates."""
        if not dates_list:
            return ["Please select at least one date."]

        # Create a map of dates to their column numbers for fast lookup
        date_to_col = {sheet.cell(row=2, column=col).value: col for col in range(4, sheet.max_column + 1)}
        
        report_lines = []
        for date_str in dates_list:
            date_col = date_to_col.get(date_str)
            if date_col is None:
                report_lines.append(f"Date '{date_str}' not found.")
                continue

            present_count, absent_count = 0, 0
            hours = sheet.cell(row=3, column=date_col).value or "N/A"
            for row in range(5, count_student_rows(sheet) + 5):
                status = sheet.cell(row=row, column=date_col).value
                if status == 'P':
                    present_count += 1
                elif status == 'A':
                    absent_count += 1
            
            total = present_count + absent_count
            report_lines.append(
                f"Subject: {sheet.title}\n"
                f"Report for {date_str}(Session Hours: {hours}):\n"
                f"  - Present: {present_count} / {total}\n"
                f"  - Absent: {absent_count} / {total}"
            )
        return report_lines

    def get_report_by_name(self, sheet, names_list):
        """Generates a summary including attendance, marks, and final result."""
        summary_cols = {}
        perc_col = self._find_percentage_col(sheet)
        if perc_col:
            summary_cols["PERCENTAGE"] = perc_col
            summary_cols["HOURS ABSENT"] = perc_col - 1
            summary_cols["HOURS PRESENT"] = perc_col - 2
        
        name_to_row = {str(sheet.cell(row, 2).value).upper(): row for row in range(5, count_student_rows(sheet) + 5)}
        
        report_lines = []
        for name in names_list:
            row_num = name_to_row.get(name.upper())
            if row_num:
                line = f"{name}: In subject ({sheet.title})"
                # Add Attendance Data
                if summary_cols:
                    hp = sheet.cell(row=row_num, column=summary_cols["HOURS PRESENT"]).value
                    ha = sheet.cell(row=row_num, column=summary_cols["HOURS ABSENT"]).value
                    perc = sheet.cell(row=row_num, column=summary_cols["PERCENTAGE"]).value
                    line += f"\n  - Hours Present: {hp}\n  - Hours Absent: {ha}\n  - Percentage: {perc}%"
                
                # Add Marks Data
                assessments = self.get_assessment_list(sheet)
                if assessments:
                    line += "\n  --- Marks ---"
                    for assessment_name in assessments:
                        col_idx = [c for c in range(1, sheet.max_column + 1) if sheet.cell(row=4, column=c).value == assessment_name][0]
                        mark = sheet.cell(row=row_num, column=col_idx).value
                        max_mark = str(sheet.cell(row=3, column=col_idx).value or '').replace('Out of: ','')
                        if mark is not None:
                            line += f"\n  - {assessment_name}: {mark}/{max_mark}"
                
                # --- NEW: Add Final Result Data ---
                final_result_col = None
                for col in range(1, sheet.max_column + 1):
                    if sheet.cell(row=4, column=col).value == "FINAL RESULT":
                        final_result_col = col
                        break
                
                line += "\n  --- Final Result ---"
                if final_result_col:
                    final_mark = sheet.cell(row=row_num, column=final_result_col).value
                    line += f"\n  - FINAL RESULT: {final_mark if final_mark is not None else 'Not Calculated'}"
                else:
                    line += "\n  - FINAL RESULT: Not Calculated"

                report_lines.append(line)
            else:
                report_lines.append(f"{name}:\n  - STUDENT NOT FOUND")
        return report_lines

    def open_detailed_report_window(self):
        """Opens the new detailed report window."""
        self.hide_status()
        if not self.wb: return self.show_status("No file loaded.", is_error=True)
        subject_name = self.subject_combo.get()
        if not subject_name: return self.show_status("Please select a subject first.", is_error=True)
        if self.detail_win is not None and self.detail_win.winfo_exists():
            return self.detail_win.focus()
        try:
            sheet = self.wb[subject_name]
            self.detail_win = DetailedReportWindow(self, sheet)
        except Exception as e:
            self.show_status(f"Could not open report window: {e}", is_error=True)

    def open_low_attendance_window(self):
        self.hide_status()
        if not self.wb: return self.show_status("No file loaded.", is_error=True)
        subject_name = self.subject_combo.get()
        if not subject_name: return self.show_status("Please select a subject.", is_error=True)
        if self.report_win and self.report_win.winfo_exists(): return self.report_win.focus()
        try:
            sheet = self.wb[subject_name]
            self.report_win = LowAttendanceWindow(self, subject_name, sheet)
        except Exception as e:
            self.show_status(f"Could not open report. Error: {e}", is_error=True)

    def get_low_attendance_students(self, sheet, threshold_percent):
        """Gets a list of students below a certain attendance percentage."""
        percentage_col = self._find_percentage_col(sheet)
        low_attendance_students = []
        for row in range(5, count_student_rows(sheet) + 5):
            name_cell = sheet.cell(row=row, column=2)
            if not name_cell.value: continue
            percent_str = str(sheet.cell(row=row, column=percentage_col).value).replace('%', '')
            try:
                percentage = float(percent_str)
                if percentage < threshold_percent:
                    low_attendance_students.append(f"{name_cell.value} ({percentage:.2f}%)")
            except (ValueError, TypeError): continue
        return low_attendance_students

    def validate_and_submit(self):
        self.hide_status()
        if not self.wb: return self.show_status("No file loaded.", is_error=True)
        subject_name = self.subject_combo.get()
        if not subject_name: return self.show_status("Field required: Subject", is_error=True)
        date_str = self.date_entry.get()
        if not date_str: return self.show_status("Field required: Date", is_error=True)
        try:
            datetime.strptime(date_str, "%d-%m-%Y")
        except ValueError: return self.show_status("Invalid date format. Use DD-MM-YYYY.", is_error=True)
        try:
            num_hours = int(self.hours_entry.get())
            if not 1 <= num_hours <= 8: return self.show_status("Hours must be 1-8.", is_error=True)
        except (ValueError, TypeError): return self.show_status("Hours must be a whole number.", is_error=True)
        
        try:
            sheet = self.wb[subject_name]
            total_students = count_student_rows(sheet)
            if total_students == 0: return self.show_status(f"No students in '{subject_name}'.", is_error=True)
        except KeyError: return self.show_status(f"Worksheet '{subject_name}' not found.", is_error=True)
        
        # --- NEW: Check if the date already exists ---
        existing_date_col = None
        for col in range(4, sheet.max_column + 2): # Check one extra column to be safe
            if sheet.cell(row=2, column=col).value == date_str:
                existing_date_col = col
                break
        
        # If the date exists, ask the user for confirmation to overwrite
        if existing_date_col:
            if not messagebox.askyesno("Confirm Overwrite", f"An entry for {date_str} already exists.\n\nDo you want to overwrite it with this new data?"):
                return # Stop the process if the user clicks "No"

        try:
            rolls_input_str = self.rolls_entry.get().strip()
            absent_rolls = []
            
            # Handle the "0 for all present" case
            if rolls_input_str == "0":
                absent_rolls = []
            elif rolls_input_str: # Process other numbers if the field is not empty
                parsed_rolls = [int(r.strip()) for r in rolls_input_str.split(',') if r.strip()]
                invalid_rolls = [r for r in parsed_rolls if r > total_students or r < 1]
                if invalid_rolls: return self.show_status(f"Invalid Rolls: {invalid_rolls} out of range (1-{total_students}).", is_error=True)
                
                if self.mode_var.get() == "absent":
                    absent_rolls = parsed_rolls
                else: # mode is "present"
                    all_students_set = set(range(1, total_students + 1))
                    present_students_set = set(parsed_rolls)
                    absent_rolls = sorted(list(all_students_set - present_students_set))

        except (ValueError, TypeError): return self.show_status("Invalid roll number format.", is_error=True)
        
        # The final confirmation message before marking
        confirm_text = "overwrite" if existing_date_col else "mark"
        if messagebox.askyesno("Confirm", f"Are you sure you want to {confirm_text} attendance for {subject_name} on {date_str} ({len(absent_rolls)} absentees)?"):
            success, message = self.mark_attendance(sheet, total_students, absent_rolls, num_hours, date_str, overwrite_col=existing_date_col)
            self.show_status(message, not success)
            if success: [w.delete(0, ctk.END) for w in [self.rolls_entry, self.hours_entry]]

    def get_all_students_in_workbook(self):
        """Scans every sheet to create a master list of all unique students."""
        master_student_set = set()
        if not self.wb:
            return []
        for sheet in self.wb.worksheets:
            students_in_sheet = self.get_student_list(sheet)
            master_student_set.update(students_in_sheet)
        return sorted(list(master_student_set))

    def get_summary_for_student_across_all_sheets(self, student_names_list):
        """
        Iterates through every sheet in the workbook and compiles a report for a list of students.
        """
        if not self.wb:
            return "No workbook loaded."
        if not student_names_list:
            return "Please select at least one student."

        final_report_parts = []
        
        # Loop through each student the user selected
        for student_name in student_names_list:
            student_report_parts = [f"Showing summary for student: {student_name.upper()}", "="*40]
            found_student = False

            # Scan every sheet for this student
            for sheet in self.wb.worksheets:
                name_to_row = {str(sheet.cell(row, 2).value or '').upper(): row for row in range(5, count_student_rows(sheet) + 5)}
                row_num = name_to_row.get(student_name.upper())
                
                if not row_num:
                    continue # Skip this sheet if student not found

                found_student = True
                subject_report = [f"\n--- SUBJECT: {sheet.title.upper()} ---"]
                
                # Get Attendance, Marks, and Final Result data for this sheet
                perc_col = self._find_percentage_col(sheet)
                if perc_col:
                    hp = sheet.cell(row=row_num, column=perc_col - 2).value
                    ha = sheet.cell(row=row_num, column=perc_col - 1).value
                    perc = sheet.cell(row=row_num, column=perc_col).value
                    subject_report.append(f"  - Percentage: {perc}% (Present: {hp}, Absent: {ha})")
                
                assessments = self.get_assessment_list(sheet)
                if assessments:
                    subject_report.append("  --- Marks ---")
                    for assessment_name in assessments:
                        col_idx = [c for c in range(1, sheet.max_column + 1) if sheet.cell(row=4, column=c).value == assessment_name][0]
                        mark = sheet.cell(row=row_num, column=col_idx).value
                        max_mark = str(sheet.cell(row=3, column=col_idx).value or '').replace('Out of: ','')
                        if mark is not None:
                            subject_report.append(f"  - {assessment_name}: {mark}/{max_mark}")
                
                final_result_col = None
                for col in range(1, sheet.max_column + 1):
                    if sheet.cell(row=4, column=col).value == "FINAL RESULT":
                        final_result_col = col
                        break
                if final_result_col:
                    final_mark = sheet.cell(row=row_num, column=final_result_col).value
                    subject_report.append(f"  - FINAL RESULT: {final_mark if final_mark is not None else 'N/A'}")
                
                student_report_parts.append("\n".join(subject_report))
            
            if found_student:
                final_report_parts.append("\n".join(student_report_parts))

        return "\n\n".join(final_report_parts)

    def mark_attendance(self, sheet, total_students, absent_list, num_hours, attendance_date, overwrite_col=None):
        """Final, robust logic for marking attendance with smart column insertion."""
        try:
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # --- 1. Find the current location of the summary block ---
            # We use "HOURS PRESENT" in row 4 as a reliable anchor.
            summary_start_col = None
            for col in range(1, sheet.max_column + 2):
                if sheet.cell(row=4, column=col).value == "HOURS PRESENT":
                    summary_start_col = col - 1 # The block starts with TOTAL HOURS
                    break
            
            if not summary_start_col:
                return False, "Could not find the summary block headers. Please check the sheet format."

            # --- 2. Determine which column to write attendance to ---
            attendance_col = 0
            if overwrite_col:
                attendance_col = overwrite_col
            else:
                attendance_col = 4
                while sheet.cell(row=2, column=attendance_col).value is not None:
                    attendance_col += 1
                
                # --- THIS IS THE "SMART EXPANSION" ---
                # If the next entry would be too close to the summary, insert new columns
                if attendance_col >= summary_start_col - 1:
                    sheet.insert_cols(idx=summary_start_col, amount=10)
                    # After inserting, the summary block has moved. We will find it again.
            
            # --- 3. Write the new attendance data ---
            sheet.cell(row=2, column=attendance_col).value = attendance_date
            sheet.cell(row=3, column=attendance_col).value = num_hours
            for i in range(5, total_students + 5):
                cell = sheet.cell(row=i, column=attendance_col)
                if sheet.cell(i, 1).value in absent_list:
                    cell.value, cell.fill = 'A', red_fill
                else:
                    cell.value, cell.fill = 'P', green_fill

            # --- 4. Recalculate and update the summary block in its current location ---
            # Find the summary columns again, as they might have moved
            current_summary_cols = {}
            for col in range(1, sheet.max_column + 1):
                header = sheet.cell(row=4, column=col).value
                if header in ["TOTAL HOURS", "HOURS PRESENT", "HOURS ABSENT", "PERCENTAGE"]:
                    current_summary_cols[header] = col
            
            if len(current_summary_cols) < 4:
                 return False, "Summary headers are missing after update."

            for row in range(5, total_students + 5):
                total_hours, present_hours = 0, 0
                
                # Loop through all attendance columns (up to the summary block)
                for col in range(4, current_summary_cols['TOTAL HOURS']): 
                    session_hours_val = sheet.cell(row=3, column=col).value
                    if session_hours_val:
                        try:
                            session_hours = int(session_hours_val)
                            total_hours += session_hours
                            if sheet.cell(row=row, column=col).value == 'P':
                                present_hours += session_hours
                        except (ValueError, TypeError): continue
                
                absent_hours = total_hours - present_hours
                percentage = (present_hours / total_hours * 100) if total_hours > 0 else 0
                
                # Write updated values to the dynamically found summary columns
                sheet.cell(row=row, column=current_summary_cols['TOTAL HOURS']).value = total_hours
                sheet.cell(row=row, column=current_summary_cols['HOURS PRESENT']).value = present_hours
                sheet.cell(row=row, column=current_summary_cols['HOURS ABSENT']).value = absent_hours
                sheet.cell(row=row, column=current_summary_cols['PERCENTAGE']).value = f"{percentage:.2f}"

            self.apply_standard_styles(sheet, total_students)
            self.wb.save(os.path.join(USER_DATA_PATH, self.current_filename))
            return True, "Attendance marked and summary updated!"
        except PermissionError: return False, f"Could not save. '{self.current_filename}' is open."
        except Exception as e: return False, f"An error occurred: {e}"

    def get_complex_rolls(self, sheet):
        """Gets a list of all complex roll numbers from column C."""
        return [str(sheet.cell(row=row, column=3).value) for row in range(5, count_student_rows(sheet) + 5) if sheet.cell(row=row, column=3).value]

    def open_live_session_window(self):
        """Opens the new Live OTP Session window."""
        self.hide_status()
        if not self.wb: return self.show_status("No file loaded.", is_error=True)
        subject_name = self.subject_combo.get()
        if not subject_name: return self.show_status("Please select a subject first.", is_error=True)
        try:
            sheet = self.wb[subject_name]
            # Open as a non-singleton window each time
            LiveSessionWindow(self, sheet)
        except Exception as e:
            self.show_status(f"Could not open Live Session: {e}", is_error=True)

if __name__ == "__main__":
    app = AttendanceApp()
    app.mainloop()