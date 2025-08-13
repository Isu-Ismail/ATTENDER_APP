import customtkinter as ctk
from tkinter import messagebox
import openpyxl as xl
import os
from datetime import date, datetime

# Import from our custom modules
from config import ICON_PATH, resource_path, USER_DATA_PATH
from excel_helpers import count_student_rows

class LowAttendanceWindow(ctk.CTkToplevel):
    """Interactive window to generate low attendance reports."""
    def __init__(self, master, subject_name, sheet):
        super().__init__(master)
        self.title("Low Attendance Report")
        self.geometry("450x500")
        self.transient(master)
        self.focus()
        try:
            self.iconbitmap(resource_path(ICON_PATH))
        except: pass

        self.app = master
        self.sheet = sheet
        self.subject_name = subject_name
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)
        
        controls_frame = ctk.CTkFrame(self)
        controls_frame.grid(row=0, column=0, padx=20, pady=(20, 10), sticky="ew")
        controls_frame.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(controls_frame, text="Show students below:").grid(row=0, column=0, padx=(10, 5), pady=10)
        self.percent_entry = ctk.CTkEntry(controls_frame, placeholder_text="e.g., 75", width=60)
        self.percent_entry.grid(row=0, column=1, pady=10, sticky="w")
        self.percent_entry.insert(0, "75")
        ctk.CTkLabel(controls_frame, text="%").grid(row=0, column=2, padx=(2, 10), pady=10)
        self.report_button = ctk.CTkButton(controls_frame, text="Generate Report", command=self.generate_report)
        self.report_button.grid(row=0, column=3, padx=10, pady=10)
        
        self.error_label = ctk.CTkLabel(self, text="", text_color=("#C00000", "#FF8282"))
        self.error_label.grid(row=1, column=0, padx=20, pady=(0, 10), sticky="w")
        
        self.textbox = ctk.CTkTextbox(self, corner_radius=8, font=("", 14))
        self.textbox.grid(row=2, column=0, padx=20, pady=(0, 20), sticky="nsew")
        self.textbox.insert("1.0", "Enter a percentage above and click 'Generate Report'.")
        self.textbox.configure(state="disabled")
        self.generate_report()

    def generate_report(self):
        self.error_label.configure(text="")
        try:
            threshold = float(self.percent_entry.get())
            if not 0 <= threshold <= 100:
                self.error_label.configure(text="Error: Percentage must be between 0 and 100.")
                return
        except (ValueError, TypeError):
            self.error_label.configure(text="Error: Please enter a valid number.")
            return

        student_list = self.app.get_low_attendance_students(self.sheet, threshold)
        self.textbox.configure(state="normal")
        self.textbox.delete("1.0", "end")
        
        if student_list is None:
            report_text = f"Could not find a 'PERCENTAGE' column in {self.subject_name}."
        elif not student_list:
            report_text = f"Congratulations!\n\nNo students in {self.subject_name} below {threshold}%."
        else:
            header = f"Students in {self.subject_name} below {threshold}% attendance:\n{'-'*50}\n"
            report_text = header + "\n".join(student_list)
        
        self.textbox.insert("1.0", report_text)
        self.textbox.configure(state="disabled")

class ManageWindow(ctk.CTkToplevel):
    """Window for creating subjects and managing student lists with separate name/roll number fields."""
    def __init__(self, master):
        super().__init__(master)
        self.title("Subject & Student Management")
        self.geometry("500x700") # Increased height slightly
        self.transient(master)
        self.focus()
        try:
            self.iconbitmap(resource_path(ICON_PATH))
        except: pass

        self.app = master
        self.current_filename = self.app.current_filename 
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1) # Configure main content row to expand
        
        # --- Subject Management (No change) ---
        subject_frame = ctk.CTkFrame(self)
        subject_frame.grid(row=0, column=0, padx=20, pady=(20, 10), sticky="ew")
        subject_frame.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(subject_frame, text="Add New Subject", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, columnspan=2, padx=10, pady=(10,0), sticky="w")
        self.new_subject_entry = ctk.CTkEntry(subject_frame, placeholder_text="Enter new subject name")
        self.new_subject_entry.grid(row=1, column=0, padx=10, pady=10, sticky="ew")
        self.add_subject_button = ctk.CTkButton(subject_frame, text="Add Subject", width=120, command=self.add_subject)
        self.add_subject_button.grid(row=1, column=1, padx=(0,10), pady=10)

        # --- NEW: Frame to copy data from another subject ---
        copy_frame = ctk.CTkFrame(self)
        copy_frame.grid(row=1, column=0, padx=20, pady=10, sticky="ew")
        copy_frame.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(copy_frame, text="Copy student list from:").grid(row=0, column=0, padx=10, pady=10, sticky="w")
        self.copy_source_combo = ctk.CTkComboBox(copy_frame, state="readonly", values=[])
        self.copy_source_combo.grid(row=0, column=1, padx=5, pady=10, sticky="ew")
        self.copy_data_button = ctk.CTkButton(copy_frame, text="Copy Data", width=120, command=self.copy_student_data)
        self.copy_data_button.grid(row=0, column=2, padx=10, pady=10)
        
        # --- Student Management (Redesigned) ---
        student_frame = ctk.CTkFrame(self)
        student_frame.grid(row=2, column=0, padx=20, pady=0, sticky="nsew")
        student_frame.grid_columnconfigure(0, weight=1)
        student_frame.grid_rowconfigure(4, weight=1)
        
        controls_frame = ctk.CTkFrame(student_frame, fg_color="transparent")
        controls_frame.grid(row=0, column=0, columnspan=2, padx=10, pady=(10,0), sticky="ew")
        ctk.CTkLabel(controls_frame, text="Update Student List", font=ctk.CTkFont(weight="bold")).pack(side="left")
        generator_button = ctk.CTkButton(controls_frame, text="Generate Roll Numbers...", width=160, command=self.open_generator_dialog)
        generator_button.pack(side="right")
        
        ctk.CTkLabel(student_frame, text="Max Students in Class:").grid(row=1, column=0, padx=10, pady=(10,0), sticky="w")
        self.max_students_entry = ctk.CTkEntry(student_frame, placeholder_text="e.g., 65")
        self.max_students_entry.grid(row=1, column=1, padx=10, pady=(10,0), sticky="ew")
        
        ctk.CTkLabel(student_frame, text="Select Subject to Manage:").grid(row=2, column=0, padx=10, pady=(10,0), sticky="w")
        self.subject_select_combo = ctk.CTkComboBox(student_frame, state="readonly", command=self.load_student_data)
        self.subject_select_combo.grid(row=2, column=1, padx=10, pady=(10,0), sticky="ew")
        
        textbox_frame = ctk.CTkFrame(student_frame, fg_color="transparent")
        textbox_frame.grid(row=4, column=0, columnspan=2, padx=10, pady=10, sticky="nsew")
        textbox_frame.grid_columnconfigure(0, weight=2)
        textbox_frame.grid_columnconfigure(1, weight=1)
        textbox_frame.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(textbox_frame, text="Student Names (Column B)").grid(row=3, column=0)
        self.names_textbox = ctk.CTkTextbox(textbox_frame, font=("", 14))
        self.names_textbox.grid(row=4, column=0, padx=(0, 5), sticky="nsew")

        ctk.CTkLabel(textbox_frame, text="Roll Numbers (Column C)").grid(row=3, column=1)
        self.rolls_textbox = ctk.CTkTextbox(textbox_frame, font=("", 14))
        self.rolls_textbox.grid(row=4, column=1, padx=(5, 0), sticky="nsew")

        self.update_students_button = ctk.CTkButton(self, text="Save Student List for Selected Subject", command=self.update_students)
        self.update_students_button.grid(row=3, column=0, padx=20, pady=20, sticky="ew")
        
        self.refresh_subject_list()
    
    def copy_student_data(self):
        """Copies student data from a source subject to the current textboxes."""
        source_subject = self.copy_source_combo.get()
        target_subject = self.subject_select_combo.get()

        if not source_subject or "No subjects" in source_subject:
            messagebox.showerror("Error", "Please select a source subject to copy from.", parent=self)
            return
        if not target_subject or "No subjects" in target_subject:
            messagebox.showerror("Error", "Please select a destination subject to copy to.", parent=self)
            return
        if source_subject == target_subject:
            messagebox.showwarning("Warning", "Source and destination subjects are the same.", parent=self)
            return
            
        try:
            source_sheet = self.app.wb[source_subject]
            num_students = count_student_rows(source_sheet)
            
            # Read names and rolls from the source sheet
            names = [str(source_sheet.cell(row=row, column=2).value or '') for row in range(5, num_students + 5)]
            rolls = [str(source_sheet.cell(row=row, column=3).value or '') for row in range(5, num_students + 5)]
            
            # Populate the textboxes
            self.names_textbox.delete("1.0", "end")
            self.names_textbox.insert("1.0", "\n".join(names))
            self.rolls_textbox.delete("1.0", "end")
            self.rolls_textbox.insert("1.0", "\n".join(rolls))
            
            # Update the max students count
            self.max_students_entry.delete(0, "end")
            self.max_students_entry.insert(0, str(num_students))

            messagebox.showinfo("Success", f"Copied {num_students} students from '{source_subject}'.\nClick 'Save Student List' to confirm.", parent=self)
        except Exception as e:
            messagebox.showerror("Error", f"Could not copy data: {e}", parent=self)

    def refresh_subject_list(self):
        """Reloads the list of subjects from the workbook into BOTH dropdowns."""
        if self.app.wb:
            subjects = self.app.wb.sheetnames
            # Configure both dropdowns with the same list of subjects
            self.subject_select_combo.configure(values=subjects)
            self.copy_source_combo.configure(values=subjects)
            
            if subjects:
                self.subject_select_combo.set(subjects[0])
                self.copy_source_combo.set(subjects[0])
                self.load_student_data(subjects[0])
            else:
                self.subject_select_combo.set("No subjects created yet")
                self.copy_source_combo.set("No subjects available")
        else:
            self.subject_select_combo.set("No file loaded")
            self.copy_source_combo.set("No file loaded")

    def load_student_data(self, selected_subject):
        """Loads existing student names and roll numbers into the separate textboxes."""
        # Clear both textboxes
        self.names_textbox.delete("1.0", "end")
        self.rolls_textbox.delete("1.0", "end")
        if not self.app.wb: return
        
        try:
            sheet = self.app.wb[selected_subject]
            # Get data from Column B (Names) and C (Roll Numbers)
            names = [str(sheet.cell(row=row, column=2).value or '') for row in range(5, count_student_rows(sheet) + 5)]
            rolls = [str(sheet.cell(row=row, column=3).value or '') for row in range(5, count_student_rows(sheet) + 5)]
            
            # Insert data into the correct textboxes
            self.names_textbox.insert("1.0", "\n".join(names))
            self.rolls_textbox.insert("1.0", "\n".join(rolls))
        except Exception as e:
            print(f"Error loading student data: {e}")

    # This function is new
    def populate_from_generator(self, rolls_list):
        """Clears the ROLLS textbox and fills it with the generated roll numbers."""
        self.rolls_textbox.delete("1.0", "end")
        self.rolls_textbox.insert("1.0", "\n".join(rolls_list))

    # This function is new
    def open_generator_dialog(self):
        """Opens the Roll Number Generator dialog window."""
        RollGeneratorDialog(self)

    # This function is updated
    def update_students(self):
        """Saves student data from the two separate textboxes."""
        selected_subject = self.subject_select_combo.get()
        if not selected_subject or "No subjects" in selected_subject: return messagebox.showerror("Error", "Please select a valid subject.", parent=self)
        try:
            max_students = int(self.max_students_entry.get())
        except (ValueError, TypeError): return messagebox.showerror("Error", "Please enter a valid number for 'Max Students'.", parent=self)
        
        student_names = [name.strip().upper() for name in self.names_textbox.get("1.0", "end").strip().splitlines() if name.strip()]
        student_rolls = [roll.strip() for roll in self.rolls_textbox.get("1.0", "end").strip().splitlines() if roll.strip()]
        
        if len(student_names) != len(student_rolls):
            return messagebox.showerror("Error", f"Mismatch: There are {len(student_names)} names but {len(student_rolls)} roll numbers. The lists must match.", parent=self)
        if len(student_names) > max_students: return messagebox.showerror("Error", f"Student count ({len(student_names)}) exceeds maximum ({max_students}).", parent=self)
        if not messagebox.askyesno("Confirm", f"Overwrite student list for '{selected_subject}' with {len(student_names)} students?", parent=self): return
        
        try:
            sheet = self.app.wb[selected_subject]
            # Clear old student data from columns A, B, and C
            for row in range(5, sheet.max_row + 5):
                for col in range(1, 4): sheet.cell(row=row, column=col).value = None
            
            # Write new student data from the two textboxes
            for i in range(len(student_names)):
                sheet.cell(row=i+5, column=1).value = i + 1              # Simple Roll No.
                sheet.cell(row=i+5, column=2).value = student_names[i]   # Name
                sheet.cell(row=i+5, column=3).value = student_rolls[i]   # Complex Roll No.
            
            self.app.apply_standard_styles(sheet, len(student_names))
            self.app.wb.save(os.path.join(USER_DATA_PATH, self.current_filename))
            messagebox.showinfo("Success", f"Student list for '{selected_subject}' updated.", parent=self)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update students: {e}", parent=self)
            
    # The other functions (add_subject, etc.) remain the same
    def add_subject(self):
        new_name = self.new_subject_entry.get().strip()
        if not new_name: return messagebox.showerror("Error", "Subject name cannot be empty.", parent=self)
        if self.app.wb is None:
            self.app.wb = xl.Workbook()
            self.app.wb.remove(self.app.wb.active)
        if new_name in self.app.wb.sheetnames: return messagebox.showerror("Error", f"A subject named '{new_name}' already exists.", parent=self)
        
        new_sheet = self.app.wb.create_sheet(title=new_name)
        self.app.format_new_sheet(new_sheet)
        try:
            self.app.wb.save(os.path.join(USER_DATA_PATH, self.current_filename))
            messagebox.showinfo("Success", f"Subject '{new_name}' was created.", parent=self)
            self.new_subject_entry.delete(0, "end")
            self.refresh_subject_list()
            self.app.update_main_subject_list()
        except Exception as e:
            messagebox.showerror("Error", f"Could not save file: {e}", parent=self)

class DetailedReportWindow(ctk.CTkToplevel):
    """A new window with tabs for generating reports by date or by name."""
    def __init__(self, master, sheet):
        super().__init__(master)
        self.title("Detailed Report Generator")
        self.geometry("500x600")
        self.transient(master)
        self.focus()
        try:
            self.iconbitmap(resource_path(ICON_PATH))
        except: pass

        self.app = master
        self.sheet = sheet

        # --- Create a Tab View to switch between report types ---
        self.tab_view = ctk.CTkTabview(self, width=480)
        self.tab_view.pack(padx=20, pady=40, fill="both", expand=True)
        self.tab_view.add("By Date")
        self.tab_view.add("By Name")
        self.tab_view.add("Student Summary")
        
        # --- Populate the "By Date" Tab ---
        self.setup_date_tab()
        
        # --- Populate the "By Name" Tab ---
        self.setup_name_tab()
        self.setup_student_summary_tab()

    def setup_date_tab(self):
        """Creates the widgets for the 'By Date' report tab with a collapsible list."""
        date_tab = self.tab_view.tab("By Date")
        
        top_frame = ctk.CTkFrame(date_tab)
        top_frame.pack(padx=10, pady=10, fill="x")

        ctk.CTkLabel(top_frame, text="Select dates from the list:").pack(side="left", anchor="w", padx=10, pady=5)

        self.date_toggle_button = ctk.CTkButton(top_frame, text="Hide Date List", width=140, command=self.toggle_date_list)
        self.date_toggle_button.pack(side="right", anchor="e", padx=10, pady=5)
        
        # This frame can now be hidden/shown
        self.date_checklist_frame = ctk.CTkScrollableFrame(date_tab, label_text="Select Dates")
        self.date_checklist_frame.pack(padx=10, pady=10, fill="both", expand=True)

        self.date_checkboxes = {}
        all_dates = self.app.get_all_dates_from_sheet(self.sheet)
        for date_str in all_dates:
            var = ctk.StringVar(value="off")
            cb = ctk.CTkCheckBox(self.date_checklist_frame, text=date_str, variable=var, onvalue="on", offvalue="off")
            cb.pack(anchor="w", padx=10, pady=2)
            self.date_checkboxes[date_str] = var

        generate_btn = ctk.CTkButton(date_tab, text="Generate Date Report", command=self.generate_date_report)
        generate_btn.pack(padx=10, pady=10, fill="x")
        
        self.date_results_textbox = ctk.CTkTextbox(date_tab, corner_radius=8, font=("", 14))
        self.date_results_textbox.pack(padx=10, pady=10, fill="both", expand=True)
        self.date_results_textbox.insert("1.0", "Select one or more dates and click 'Generate Report'.")
        self.date_results_textbox.configure(state="disabled")

    # Add this new function inside the DetailedReportWindow class
    def toggle_student_list(self):
        """Shows or hides the student checklist frame."""
        if self.checklist_frame.winfo_viewable():
            self.checklist_frame.pack_forget()
            self.toggle_button.configure(text="Show Student List")
        else:
            self.checklist_frame.pack(padx=10, pady=5, fill="both", expand=True)
            self.toggle_button.configure(text="Hide Student List")
    
    def toggle_date_list(self):
        """Shows or hides the date checklist frame."""
        if self.date_checklist_frame.winfo_viewable():
            self.date_checklist_frame.pack_forget()
            self.date_toggle_button.configure(text="Show Date List")
        else:
            self.date_checklist_frame.pack(padx=10, pady=5, fill="both", expand=True)
            self.date_toggle_button.configure(text="Hide Date List")

    # Replace your old setup_name_tab function with this one
    def setup_name_tab(self):
        """Creates the widgets for the 'By Name' report tab with a collapsible list."""
        name_tab = self.tab_view.tab("By Name")
        
        top_frame = ctk.CTkFrame(name_tab)
        top_frame.pack(padx=10, pady=10, fill="x")
        
        ctk.CTkLabel(top_frame, text="Select students from the list:").pack(side="left", anchor="w", padx=10, pady=5)

        # This button will control the visibility of the checklist
        self.toggle_button = ctk.CTkButton(top_frame, text="Hide Student List", width=140, command=self.toggle_student_list)
        self.toggle_button.pack(side="right", anchor="e", padx=10, pady=5)
        
        # Create a scrollable frame for the student checklist
        self.checklist_frame = ctk.CTkScrollableFrame(name_tab, label_text="Student List")
        self.checklist_frame.pack(padx=10, pady=5, fill="both", expand=True)

        self.student_checkboxes = {}
        student_names = self.app.get_student_list(self.sheet)
        for name in student_names:
            var = ctk.StringVar(value="off")
            cb = ctk.CTkCheckBox(self.checklist_frame, text=name, variable=var, onvalue="on", offvalue="off")
            cb.pack(anchor="w", padx=10, pady=2)
            self.student_checkboxes[name] = var

        generate_btn = ctk.CTkButton(name_tab, text="Generate Name Report", command=self.generate_name_report)
        generate_btn.pack(padx=10, pady=10, fill="x")
        
        self.name_results_textbox = ctk.CTkTextbox(name_tab, corner_radius=8, font=("", 14))
        self.name_results_textbox.pack(padx=10, pady=10, fill="both", expand=True)
        self.name_results_textbox.insert("1.0", "Select students and click 'Generate Report'.")
        self.name_results_textbox.configure(state="disabled")

    def generate_date_report(self):
        """Gathers selected dates and generates the report."""
        selected_dates = [date_str for date_str, var in self.date_checkboxes.items() if var.get() == "on"]
        
        report_lines = self.app.get_report_by_date(self.sheet, selected_dates)
        report_text = "\n\n".join(report_lines)
            
        self.date_results_textbox.configure(state="normal")
        self.date_results_textbox.delete("1.0", "end")
        self.date_results_textbox.insert("1.0", report_text)
        self.date_results_textbox.configure(state="disabled")

    def generate_name_report(self):
        selected_names = [name for name, var in self.student_checkboxes.items() if var.get() == "on"]
        
        if not selected_names:
            report_text = "Please select at least one student from the checklist."
        else:
            report_lines = self.app.get_report_by_name(self.sheet, selected_names)
            report_text = "\n\n".join(report_lines)
            
        self.name_results_textbox.configure(state="normal")
        self.name_results_textbox.delete("1.0", "end")
        self.name_results_textbox.insert("1.0", report_text)
        self.name_results_textbox.configure(state="disabled")

    def setup_student_summary_tab(self):
        """Creates the UI for the cross-subject student summary report using checkboxes."""
        summary_tab = self.tab_view.tab("Student Summary")

        top_frame = ctk.CTkFrame(summary_tab)
        top_frame.pack(padx=10, pady=10, fill="x")
        ctk.CTkLabel(top_frame, text="Select students for a full report:").pack(side="left", anchor="w", padx=10, pady=5)
        
        self.summary_toggle_button = ctk.CTkButton(top_frame, text="Hide Student List", width=140, command=self.toggle_student_summary_list)
        self.summary_toggle_button.pack(side="right", anchor="e", padx=10, pady=5)

        self.summary_checklist_frame = ctk.CTkScrollableFrame(summary_tab, label_text="Master Student List")
        self.summary_checklist_frame.pack(padx=10, pady=5, fill="both", expand=True)

        self.summary_checkboxes = {} # Use checkboxes for multi-select
        all_students = self.app.get_all_students_in_workbook()
        for name in all_students:
            var = ctk.StringVar(value="off")
            cb = ctk.CTkCheckBox(self.summary_checklist_frame, text=name, variable=var, onvalue="on", offvalue="off")
            cb.pack(anchor="w", padx=10, pady=2)
            self.summary_checkboxes[name] = var

        generate_btn = ctk.CTkButton(summary_tab, text="Generate Student Summary", command=self.generate_student_summary_report)
        generate_btn.pack(padx=10, pady=10, fill="x")
        
        self.summary_results_textbox = ctk.CTkTextbox(summary_tab, corner_radius=8, font=("", 14))
        self.summary_results_textbox.pack(padx=10, pady=10, fill="both", expand=True)
        self.summary_results_textbox.insert("1.0", "Select one or more students and click 'Generate'.")
        self.summary_results_textbox.configure(state="disabled")

    def generate_student_summary_report(self):
        """Gathers data for multiple students across all sheets."""
        selected_students = [name for name, var in self.summary_checkboxes.items() if var.get() == "on"]
        
        report_text = self.app.get_summary_for_student_across_all_sheets(selected_students)
            
        self.summary_results_textbox.configure(state="normal")
        self.summary_results_textbox.delete("1.0", "end")
        self.summary_results_textbox.insert("1.0", report_text)
        self.summary_results_textbox.configure(state="disabled")

    def toggle_student_summary_list(self):
        """Shows or hides the master student list frame."""
        if self.summary_checklist_frame.winfo_viewable():
            self.summary_checklist_frame.pack_forget()
            self.summary_toggle_button.configure(text="Show Student List")
        else:
            self.summary_checklist_frame.pack(padx=10, pady=5, fill="both", expand=True)
            self.summary_toggle_button.configure(text="Hide Student List")

class RollGeneratorDialog(ctk.CTkToplevel):
    """A dialog for generating complex roll numbers based on rules."""
    def __init__(self, master):
        super().__init__(master)
        self.title("Roll Number Generator")
        self.geometry("400x300")
        self.transient(master)
        self.focus()

        self.manage_window = master # Reference to the ManageWindow that opened it

        self.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(self, text="Enter Generation Rules", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=20, pady=(20,10))
        
        main_frame = ctk.CTkFrame(self)
        main_frame.grid(row=1, column=0, padx=20, pady=10, sticky="ew")
        main_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(main_frame, text="Prefixes (comma-sep):").grid(row=0, column=0, padx=10, pady=5, sticky="w")
        self.prefixes_entry = ctk.CTkEntry(main_frame, placeholder_text="e.g., 20235070, 20235073")
        self.prefixes_entry.grid(row=0, column=1, padx=10, pady=5, sticky="ew")
        
        ctk.CTkLabel(main_frame, text="Ranges (comma-sep):").grid(row=1, column=0, padx=10, pady=5, sticky="w")
        self.ranges_entry = ctk.CTkEntry(main_frame, placeholder_text="e.g., 1-53, 1-12")
        self.ranges_entry.grid(row=1, column=1, padx=10, pady=5, sticky="ew")
        
        ctk.CTkLabel(main_frame, text="Exclusions (comma-sep):").grid(row=2, column=0, padx=10, pady=5, sticky="w")
        self.exclusions_entry = ctk.CTkEntry(main_frame, placeholder_text="e.g., 2023507034")
        self.exclusions_entry.grid(row=2, column=1, padx=10, pady=5, sticky="ew")

        generate_button = ctk.CTkButton(self, text="Generate and Paste", command=self.generate_and_paste)
        generate_button.grid(row=2, column=0, padx=20, pady=20)

    def generate_and_paste(self):
        """Parses inputs, generates roll numbers, and pastes them into the ManageWindow."""
        try:
            prefixes = [p.strip() for p in self.prefixes_entry.get().split(',')]
            ranges = [r.strip() for r in self.ranges_entry.get().split(',')]
            exclusions = {e.strip() for e in self.exclusions_entry.get().split(',') if e.strip()}

            if len(prefixes) != len(ranges):
                messagebox.showerror("Error", "The number of prefixes must match the number of ranges.", parent=self)
                return

            generated_rolls = []
            for prefix, r_str in zip(prefixes, ranges):
                start, end = map(int, r_str.split('-'))
                for i in range(start, end + 1):
                    # Format with leading zero if needed (e.g., 1 -> 01, 10 -> 10)
                    roll_suffix = f"{i:02d}"
                    full_roll = f"{prefix}{roll_suffix}"
                    if full_roll not in exclusions:
                        generated_rolls.append(full_roll)
            
            # Pass the list back to the ManageWindow
            self.manage_window.populate_from_generator(generated_rolls)
            self.destroy() # Close the generator window

        except Exception as e:
            messagebox.showerror("Error", f"Invalid input format. Please check your entries.\nDetails: {e}", parent=self)

# In ui_windows.py
class BulkEntryWindow(ctk.CTkToplevel):
    """A window for entering multiple attendance records at once."""
    def __init__(self, master, sheet):
        super().__init__(master)
        self.title("Bulk Attendance Entry")
        self.geometry("600x650")
        self.transient(master)
        self.focus()
        try:
            self.iconbitmap(resource_path(ICON_PATH))
        except: pass

        self.app = master
        self.sheet = sheet
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)
        
        ctk.CTkLabel(self, text="Paste or type attendance data below.", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=20, pady=(20, 5), sticky="w")
        ctk.CTkLabel(self, text="Format: DATE:HOURS:ABSENTEE_ROLLS (e.g., 08-08-2025:2:1,3,5)", text_color="gray").grid(row=1, column=0, padx=20, pady=(0, 10), sticky="w")
        
        self.input_textbox = ctk.CTkTextbox(self, font=("", 14), height=150)
        self.input_textbox.grid(row=2, column=0, padx=20, pady=5, sticky="nsew")

        self.process_button = ctk.CTkButton(self, text="Process Bulk Entries", command=self.process_entries)
        self.process_button.grid(row=3, column=0, padx=20, pady=10, sticky="ew")

        self.results_textbox = ctk.CTkTextbox(self, corner_radius=8, font=("", 12), state="disabled")
        self.results_textbox.grid(row=4, column=0, padx=20, pady=(5, 20), sticky="nsew")

    def log_message(self, message):
        """Adds a message to the results log textbox."""
        self.results_textbox.configure(state="normal")
        self.results_textbox.insert("end", message + "\n")
        self.results_textbox.configure(state="disabled")
        self.results_textbox.see("end")
        self.update_idletasks()

    def _parse_date(self, date_str):
        """Tries to parse a date string using multiple common formats."""
        # A list of date formats to try
        formats_to_try = [
            "%d-%m-%Y",  # 08-08-2025
            "%d/%m/%Y",  # 08/08/2025
            "%d-%m-%y",  # 08-08-25
            "%d/%m/%y",  # 08/08/25
        ]
        for fmt in formats_to_try:
            try:
                # If parsing is successful, return the date formatted in our standard way
                parsed_date = datetime.strptime(date_str, fmt)
                return parsed_date.strftime("%d-%m-%Y")
            except ValueError:
                continue # Try the next format
        
        # If all formats fail, return None
        return None

    def process_entries(self):
        """Validates and processes each line from the input textbox."""
        self.process_button.configure(state="disabled")
        self.results_textbox.configure(state="normal")
        self.results_textbox.delete("1.0", "end")
        self.results_textbox.configure(state="disabled")

        self.log_message("--- Starting bulk processing ---")
        
        all_lines = self.input_textbox.get("1.0", "end").strip().splitlines()
        total_students = count_student_rows(self.sheet)

        for i, line in enumerate(all_lines):
            line = line.strip()
            if not line: continue

            self.log_message(f"\nProcessing line {i+1}: '{line}'")
            try:
                parts = line.split(':')
                if len(parts) != 3:
                    self.log_message("  -> ERROR: Invalid format. Must be DATE:HOURS:ROLLS.")
                    continue
                
                date_input, hours_str, rolls_str = [p.strip() for p in parts]
                
                # --- NEW: Use the smart date parser ---
                date_str = self._parse_date(date_input)
                if date_str is None:
                    self.log_message(f"  -> ERROR: Invalid date format for '{date_input}'.")
                    continue
                
                num_hours = int(hours_str)
                if not 1 <= num_hours <= 8:
                    self.log_message("  -> ERROR: Hours must be between 1 and 8.")
                    continue

                parsed_rolls = [int(r.strip()) for r in rolls_str.split(',') if r.strip()] if rolls_str else []
                invalid_rolls = [r for r in parsed_rolls if r > total_students or r < 1]
                if invalid_rolls:
                    self.log_message(f"  -> ERROR: Invalid Rolls: {invalid_rolls} out of range (1-{total_students}).")
                    continue
                
                existing_date_col = None
                for col in range(4, self.sheet.max_column + 2):
                    if self.sheet.cell(row=2, column=col).value == date_str:
                        existing_date_col = col
                        break
                
                if existing_date_col:
                    if not messagebox.askyesno("Confirm Overwrite", f"An entry for {date_str} already exists.\n\nDo you want to overwrite it?", parent=self):
                        self.log_message(f"  -> SKIPPED: User chose not to overwrite date {date_str}.")
                        continue

                success, message = self.app.mark_attendance(self.sheet, total_students, parsed_rolls, num_hours, date_str, overwrite_col=existing_date_col)
                self.log_message(f"  -> STATUS: {message}")

            except Exception as e:
                self.log_message(f"  -> FAILED: An unexpected error occurred. ({e})")
        
        self.log_message("\n--- Bulk processing complete! ---")
        self.process_button.configure(state="normal")

    def log_message(self, message):
        """Adds a message to the results log textbox."""
        self.results_textbox.configure(state="normal")
        self.results_textbox.insert("end", message + "\n")
        self.results_textbox.configure(state="disabled")
        self.results_textbox.see("end") # Auto-scroll to the bottom
        self.update_idletasks() # Force UI to update

    def process_entries(self):
        """Validates and processes each line from the input textbox with detailed logging."""
        self.process_button.configure(state="disabled")
        self.results_textbox.configure(state="normal")
        self.results_textbox.delete("1.0", "end")
        self.results_textbox.configure(state="disabled")

        self.log_message("--- Starting bulk processing ---")
        
        all_lines = self.input_textbox.get("1.0", "end").strip().splitlines()
        total_students = count_student_rows(self.sheet)
        
        for i, line in enumerate(all_lines):
            line = line.strip()
            if not line: continue

            self.log_message(f"\nProcessing line {i+1}: '{line}'")

            # 1. Validate the overall format (must have 3 parts separated by ':')
            parts = line.split(':')
            if len(parts) != 3:
                self.log_message("  -> ERROR: Invalid format. Expected DATE:HOURS:ROLLS.")
                continue
            
            date_input, hours_str, rolls_str = [p.strip() for p in parts]

            # 2. Validate the Date
            date_str = self._parse_date(date_input)
            if date_str is None:
                self.log_message(f"  -> ERROR: Invalid date format '{date_input}'. Use DD-MM-YYYY or similar.")
                continue

            # 3. Validate the Hours
            try:
                num_hours = int(hours_str)
                if not 1 <= num_hours <= 8:
                    self.log_message(f"  -> ERROR: Hours '{num_hours}' must be between 1 and 8.")
                    continue
            except (ValueError, TypeError):
                self.log_message(f"  -> ERROR: Hours '{hours_str}' is not a valid number.")
                continue

            # 4. Validate the Roll Numbers
            try:
                # Handle the "0" for all present case specifically
                if rolls_str == "0":
                    parsed_rolls = []
                else:
                    parsed_rolls = [int(r.strip()) for r in rolls_str.split(',') if r.strip()] if rolls_str else []
                
                invalid_rolls = [r for r in parsed_rolls if r > total_students or r < 1]
                if invalid_rolls:
                    self.log_message(f"  -> ERROR: Invalid Rolls {invalid_rolls} (out of range 1-{total_students}).")
                    continue
            except (ValueError, TypeError):
                self.log_message(f"  -> ERROR: Roll numbers '{rolls_str}' contain non-numeric characters.")
                continue

            # 5. Check for existing date and ask to overwrite
            existing_date_col = None
            for col in range(4, self.sheet.max_column + 2):
                if self.sheet.cell(row=2, column=col).value == date_str:
                    existing_date_col = col
                    break
            
            if existing_date_col:
                if not messagebox.askyesno("Confirm Overwrite", f"An entry for {date_str} (from line {i+1}) already exists.\n\nDo you want to overwrite it?", parent=self):
                    self.log_message(f"  -> SKIPPED: User chose not to overwrite date {date_str}.")
                    continue

            # If all validations pass, call the main mark_attendance function
            success, message = self.app.mark_attendance(self.sheet, total_students, parsed_rolls, num_hours, date_str, overwrite_col=existing_date_col)
            self.log_message(f"  -> STATUS: {message}")
        
        self.log_message("\n--- Bulk processing complete! ---")
        self.process_button.configure(state="normal")

# Add these three classes to ui_windows.py
class AddAssessmentDialog(ctk.CTkToplevel):
    """A dialog to get the name and max marks for a new assessment."""
    def __init__(self, master):
        super().__init__(master)
        self.title("Add New Assessment")
        self.geometry("350x200")
        self.transient(master)
        self.focus()
        self.lift()

        self.result = None

        self.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(self, text="Assessment Name:").grid(row=0, column=0, padx=20, pady=(20,5), sticky="w")
        self.name_entry = ctk.CTkEntry(self, placeholder_text="e.g., Midterm Exam")
        self.name_entry.grid(row=0, column=1, padx=20, pady=(20,5), sticky="ew")

        ctk.CTkLabel(self, text="Maximum Marks:").grid(row=1, column=0, padx=20, pady=5, sticky="w")
        self.marks_entry = ctk.CTkEntry(self, placeholder_text="e.g., 100")
        self.marks_entry.grid(row=1, column=1, padx=20, pady=5, sticky="ew")

        button_frame = ctk.CTkFrame(self, fg_color="transparent")
        button_frame.grid(row=2, column=0, columnspan=2, padx=20, pady=20)
        
        ok_button = ctk.CTkButton(button_frame, text="OK", command=self.on_ok)
        ok_button.pack(side="left", padx=10)
        
        cancel_button = ctk.CTkButton(button_frame, text="Cancel", command=self.on_cancel)
        cancel_button.pack(side="left", padx=10)
        
        self.name_entry.focus()
        self.wait_window()

    def on_ok(self):
        name = self.name_entry.get().strip()
        max_marks = self.marks_entry.get().strip()
        if name and max_marks:
            self.result = (name, max_marks)
            self.destroy()
        else:
            messagebox.showerror("Error", "Both fields are required.", parent=self)
            
    def on_cancel(self):
        self.result = None
        self.destroy()

class MarkEntryWindow(ctk.CTkToplevel):
    """The main window for entering and managing student marks."""
    def __init__(self, master, sheet):
        super().__init__(master)
        self.title("Mark Entry and Calculation")
        self.geometry("600x650")
        self.transient(master)
        self.focus()
        try:
            self.iconbitmap(resource_path(ICON_PATH))
        except: pass

        self.app = master
        self.sheet = sheet
        self.student_names = self.app.get_student_list(self.sheet)
        self.entry_widgets = []

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        # --- Top Controls ---
        top_frame = ctk.CTkFrame(self)
        top_frame.grid(row=0, column=0, columnspan=2, padx=20, pady=(20,10), sticky="ew")
        top_frame.grid_columnconfigure(1, weight=1)
        
        ctk.CTkLabel(top_frame, text="Select Assessment:", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=10, pady=10)
        self.assessment_combo = ctk.CTkComboBox(top_frame, state="readonly", command=self.load_marks_into_grid)
        self.assessment_combo.grid(row=0, column=1, padx=10, pady=10, sticky="ew")
        add_assessment_btn = ctk.CTkButton(top_frame, text="Add New Assessment...", command=self.add_new_assessment)
        add_assessment_btn.grid(row=0, column=2, padx=10, pady=10)

        # --- Main Data Entry Grid ---
        self.grid_frame = ctk.CTkScrollableFrame(self, label_text="Student Marks")
        self.grid_frame.grid(row=1, column=0, columnspan=2, padx=20, pady=5, sticky="nsew")
        self.grid_frame.grid_columnconfigure(0, weight=3)
        self.grid_frame.grid_columnconfigure(1, weight=1)
        
        # --- Bulk Entry Section ---
        bulk_frame = ctk.CTkFrame(self)
        bulk_frame.grid(row=2, column=0, columnspan=2, padx=20, pady=10, sticky="ew")
        bulk_frame.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(bulk_frame, text="For Bulk Entry: Paste marks here (one per line) and click Apply.").pack(padx=10, pady=(5,0))
        self.bulk_textbox = ctk.CTkTextbox(bulk_frame, height=100)
        self.bulk_textbox.pack(padx=10, pady=5, fill="x", expand=True)
        apply_bulk_btn = ctk.CTkButton(bulk_frame, text="Apply Bulk Marks to Grid Above", command=self.apply_bulk_marks)
        apply_bulk_btn.pack(padx=10, pady=(0,10))
        
        # --- NEW: Bottom Buttons Frame with all three buttons ---
        button_frame = ctk.CTkFrame(self)
        button_frame.grid(row=3, column=0, columnspan=2, padx=20, pady=20, sticky="ew")
        button_frame.grid_columnconfigure((0, 1, 2), weight=1) # Configure 3 columns
        
        save_btn = ctk.CTkButton(button_frame, text="Save All Marks", command=self.save_marks)
        save_btn.grid(row=0, column=0, padx=5, sticky="ew")

        converter_btn = ctk.CTkButton(button_frame, text="Mark Converter Tool", command=self.open_converter)
        converter_btn.grid(row=0, column=1, padx=5, sticky="ew")

        calc_btn = ctk.CTkButton(button_frame, text="Calculate Final Result", command=self.open_calculator)
        calc_btn.grid(row=0, column=2, padx=5, sticky="ew")
        
        self.populate_grid()
        self.refresh_assessments()

    def populate_grid(self):
        """Creates the student name labels and mark entry boxes."""
        for i, name in enumerate(self.student_names):
            label = ctk.CTkLabel(self.grid_frame, text=f"{i+1}. {name}")
            label.grid(row=i, column=0, padx=10, pady=5, sticky="w")
            entry = ctk.CTkEntry(self.grid_frame)
            entry.grid(row=i, column=1, padx=10, pady=5, sticky="ew")
            self.entry_widgets.append(entry)
            
    def refresh_assessments(self):
        assessments = self.app.get_assessment_list(self.sheet)
        self.assessment_combo.configure(values=assessments)
        if assessments:
            self.assessment_combo.set(assessments[0])
            self.load_marks_into_grid(assessments[0])
        else:
            self.assessment_combo.set("No assessments created yet")

    def load_marks_into_grid(self, assessment_name):
        """Loads existing marks from the sheet into the individual entry boxes."""
        marks = self.app.get_marks_for_assessment(self.sheet, assessment_name)
        for i, entry in enumerate(self.entry_widgets):
            entry.delete(0, "end")
            if i < len(marks):
                entry.insert(0, marks[i])
    
    def apply_bulk_marks(self):
        """Validates and then pastes marks from the bulk textbox into the grid."""
        # --- 1. Get context: the selected assessment and its max marks ---
        assessment_name = self.assessment_combo.get()
        if not assessment_name or "No assessments" in assessment_name:
            return messagebox.showerror("Error", "Please select an assessment first.", parent=self)
            
        max_mark = self.app.get_max_marks(self.sheet, assessment_name)
        if max_mark is None:
            return messagebox.showerror("Error", f"Could not determine max marks for '{assessment_name}'.", parent=self)

        # --- 2. Get and validate the raw text data ---
        marks_list_str = self.bulk_textbox.get("1.0", "end").strip().splitlines()
        if len(marks_list_str) != len(self.student_names):
            return messagebox.showerror("Error", f"Data mismatch: There are {len(self.student_names)} students but you pasted {len(marks_list_str)} marks.", parent=self)

        # --- 3. Validate each individual mark BEFORE applying ---
        for i, mark_str in enumerate(marks_list_str):
            if not mark_str.strip():
                continue # Allow empty marks
            try:
                mark_int = int(mark_str)
                if not 0 <= mark_int <= max_mark:
                    messagebox.showerror("Validation Error", f"Error on line {i+1} of your bulk entry:\n\nMark '{mark_int}' is out of range. It must be between 0 and {max_mark}.", parent=self)
                    return # Stop the entire process
            except (ValueError, TypeError):
                messagebox.showerror("Validation Error", f"Error on line {i+1} of your bulk entry:\n\n'{mark_str}' is not a valid number.", parent=self)
                return # Stop the entire process

        # --- 4. If all validations pass, apply the marks to the grid ---
        for i, entry in enumerate(self.entry_widgets):
            entry.delete(0, "end")
            if i < len(marks_list_str):
                entry.insert(0, marks_list_str[i])
        
        self.bulk_textbox.delete("1.0", "end")
        messagebox.showinfo("Success", "Bulk marks applied to the grid. Click 'Save All Marks' to make them permanent.", parent=self)
        
    # In ui_windows.py, inside the MarkEntryWindow class

    def open_converter(self):
        """Opens the Mark Converter dialog."""
        MarkConverterDialog(self)

    def open_calculator(self):
        """Opens the Final Result Calculator dialog."""
        FinalResultDialog(self)
    
    def add_new_assessment(self):
        dialog = AddAssessmentDialog(self)
        result = dialog.result
        if result:
            name, max_marks = result
            success, message = self.app.add_new_assessment_column(self.sheet, name, max_marks)
            if success:
                messagebox.showinfo("Success", message, parent=self)
                self.refresh_assessments()
            else:
                messagebox.showerror("Error", message, parent=self)

    def save_marks(self):
        assessment_name = self.assessment_combo.get()
        if not assessment_name or "No assessments" in assessment_name:
            return messagebox.showerror("Error", "Please select an assessment to save.", parent=self)
            
        marks_from_grid = [entry.get().strip() for entry in self.entry_widgets]
        
        max_mark = self.app.get_max_marks(self.sheet, assessment_name)
        if max_mark is None: return messagebox.showerror("Error", "Could not determine max marks.", parent=self)

        validated_marks = []
        for i, mark_str in enumerate(marks_from_grid):
            if not mark_str:
                validated_marks.append(None)
                continue
            try:
                mark_int = int(mark_str)
                if not 0 <= mark_int <= max_mark:
                    return messagebox.showerror("Validation Error", f"Error for student {i+1}: Mark '{mark_int}' is invalid. Must be between 0 and {max_mark}.", parent=self)
                validated_marks.append(mark_int)
            except (ValueError, TypeError):
                return messagebox.showerror("Validation Error", f"Error for student {i+1}: '{mark_str}' is not a valid number.", parent=self)

        if not messagebox.askyesno("Confirm Save", f"Save these marks for '{assessment_name}'?\nThis will overwrite any existing data.", parent=self):
            return

        success, message = self.app.save_marks(self.sheet, assessment_name, validated_marks)
        messagebox.showinfo("Status", message, parent=self)

# Placeholder classes for future implementation
class MarkConverterDialog(ctk.CTkToplevel):
    """A dialog to convert marks from one scale to another."""
    def __init__(self, master):
        super().__init__(master)
        self.title("Smart Mark Converter")
        self.geometry("400x300")
        self.transient(master)
        self.focus()
        self.lift()

        self.mark_entry_window = master
        self.app = self.mark_entry_window.app
        self.sheet = self.mark_entry_window.sheet
        
        self.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(self, text="Convert Marks", font=ctk.CTkFont(size=16, weight="bold")).grid(row=0, column=0, columnspan=2, padx=20, pady=20)

        ctk.CTkLabel(self, text="Select Assessment:").grid(row=1, column=0, padx=20, pady=5, sticky="w")
        self.assessment_combo = ctk.CTkComboBox(self, state="readonly", values=self.app.get_assessment_list(self.sheet), command=self.on_assessment_select)
        self.assessment_combo.grid(row=1, column=1, padx=20, pady=5, sticky="ew")

        ctk.CTkLabel(self, text="Marks are currently out of:").grid(row=2, column=0, padx=20, pady=5, sticky="w")
        self.from_entry = ctk.CTkEntry(self)
        self.from_entry.grid(row=2, column=1, padx=20, pady=5, sticky="ew")

        ctk.CTkLabel(self, text="Convert to out of:").grid(row=3, column=0, padx=20, pady=5, sticky="w")
        self.to_entry = ctk.CTkEntry(self, placeholder_text="e.g., 100")
        self.to_entry.grid(row=3, column=1, padx=20, pady=5, sticky="ew")

        self.convert_button = ctk.CTkButton(self, text="Convert Marks", command=self.convert)
        self.convert_button.grid(row=4, column=0, columnspan=2, padx=20, pady=20)
        
        if self.assessment_combo.get():
            self.on_assessment_select(self.assessment_combo.get())

    def on_assessment_select(self, assessment_name):
        """Auto-fills the 'from' entry when an assessment is selected."""
        max_mark = self.app.get_max_marks(self.sheet, assessment_name)
        self.from_entry.delete(0, "end")
        if max_mark is not None:
            self.from_entry.insert(0, str(max_mark))

    def convert(self):
        assessment = self.assessment_combo.get()
        try:
            from_val = int(self.from_entry.get())
            to_val = int(self.to_entry.get())
        except (ValueError, TypeError):
            return messagebox.showerror("Error", "Please enter valid numbers for the mark scales.", parent=self)
        
        if not messagebox.askyesno("Confirm", f"This will permanently convert all marks for '{assessment}' from a scale of {from_val} to {to_val}. This action cannot be undone. Continue?", parent=self):
            return
            
        success, message = self.app.convert_marks(self.sheet, assessment, from_val, to_val)
        if success:
            self.mark_entry_window.load_marks_into_grid(assessment) # Refresh the grid
            messagebox.showinfo("Success", message, parent=self)
            self.destroy()
        else:
            messagebox.showerror("Error", message, parent=self)

class FinalResultDialog(ctk.CTkToplevel):
    """A dialog to calculate a final weighted result."""
    def __init__(self, master):
        super().__init__(master)
        self.title("Final Result Calculator")
        self.geometry("450x450")
        self.transient(master)
        self.focus()
        self.lift()

        self.mark_entry_window = master
        self.app = self.mark_entry_window.app
        self.sheet = self.mark_entry_window.sheet
        self.weight_entries = {}

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(self, text="Calculate Final Weighted Result", font=ctk.CTkFont(size=16, weight="bold")).grid(row=0, column=0, padx=20, pady=20)

        scroll_frame = ctk.CTkScrollableFrame(self, label_text="Assign Percentage Weights")
        scroll_frame.grid(row=1, column=0, padx=20, pady=5, sticky="nsew")
        scroll_frame.grid_columnconfigure(0, weight=2)
        scroll_frame.grid_columnconfigure(1, weight=1)
        
        for i, assessment in enumerate(self.app.get_assessment_list(self.sheet)):
            ctk.CTkLabel(scroll_frame, text=assessment).grid(row=i, column=0, padx=10, pady=5, sticky="w")
            entry = ctk.CTkEntry(scroll_frame, placeholder_text="%")
            entry.grid(row=i, column=1, padx=10, pady=5, sticky="ew")
            self.weight_entries[assessment] = entry

        self.calc_button = ctk.CTkButton(self, text="Calculate and Add to Sheet", command=self.calculate)
        self.calc_button.grid(row=3, column=0, padx=20, pady=20)

    def calculate(self):
        weights_dict = {}
        total_weight = 0
        try:
            for assessment, entry in self.weight_entries.items():
                weight_str = entry.get()
                if weight_str:
                    weight = float(weight_str)
                    weights_dict[assessment] = weight
                    total_weight += weight
        except (ValueError, TypeError):
            return messagebox.showerror("Error", "All weights must be valid numbers.", parent=self)
            
        if not messagebox.askyesno("Confirm Weights", f"The total weight assigned is {total_weight}%. Do you want to proceed?", parent=self):
            return

        # The final column name is now fixed
        final_col_name = "FINAL RESULT"

        success, message = self.app.calculate_final_result(self.sheet, weights_dict, final_col_name)
        if success:
            self.mark_entry_window.refresh_assessments()
            messagebox.showinfo("Success", message, parent=self)
            self.destroy()
        else:
            messagebox.showerror("Error", message, parent=self)
        # ... UI and logic for the calculator ...
# #class LoadingWindow(ctk.CTkToplevel):
#     """A simple splash screen that shows while the main app is loading."""
#     # In ui_windows.py, inside the LoadingWindow class
#     def __init__(self, master):
#         super().__init__(master)
#         self.title("Loading")
        
#         width, height = 300, 150
#         screen_width = self.winfo_screenwidth()
#         screen_height = self.winfo_screenheight()
#         x = (screen_width / 2) - (width / 2)
#         y = (screen_height / 2) - (height / 2)
#         self.geometry(f'{width}x{height}+{int(x)}+{int(y)}')
#         self.overrideredirect(True)

#         self.main_frame = ctk.CTkFrame(self, corner_radius=10)
#         self.main_frame.pack(fill="both", expand=True, padx=5, pady=5)

#         self.label = ctk.CTkLabel(self.main_frame, text="Attendance Marker", font=ctk.CTkFont(size=16, weight="bold"))
#         self.label.pack(padx=20, pady=(20,10))
        
#         self.progress_bar = ctk.CTkProgressBar(self.main_frame, mode='indeterminate')
        
#         self.start_button = ctk.CTkButton(self.main_frame, text="Start Application", width=200, height=40, command=self.start_build)
#         self.start_button.pack(padx=20, pady=10, expand=True)
    
#     def start_build(self):
#         """Hides the start button, shows the progress bar, and tells the main app to build."""
#         self.start_button.pack_forget()
#         self.label.configure(text="Loading Application...")
#         self.progress_bar.pack(padx=20, pady=10, fill="x", expand=True)
#         self.progress_bar.start()
        
#         # Call the main app's setup function after a short delay
#         self.master.after(100, self.master.setup_main_application)